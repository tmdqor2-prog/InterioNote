"""
v3.5.0 ⑰ — 데이터 내보내기 API.

GET  /api/export/clients.xlsx  : 고객 목록 엑셀 내보내기
GET  /api/export/meetings.xlsx : 상담 목록 엑셀 내보내기
GET  /api/qr/{meeting_id}      : 상담 요약 QR 코드 이미지 ⑭
"""
from __future__ import annotations

import io
import json
from datetime import datetime

from fastapi import APIRouter, HTTPException
from fastapi.responses import Response, StreamingResponse

from app.db import db_cursor

router = APIRouter(tags=["export"])


# ─── ⑰ 엑셀 내보내기 ─────────────────────────────────────────────────────────

@router.get("/api/export/clients.xlsx")
def export_clients():
    """고객 전체 목록을 xlsx 로 내보내기."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl 라이브러리가 없습니다.")

    with db_cursor() as cur:
        clients = cur.execute(
            """
            SELECT c.id, c.name, c.descriptor, c.phone, c.email, c.visit_source,
                   c.stage, c.first_met_at, c.last_meeting_at,
                   (SELECT COUNT(*) FROM meetings WHERE client_id = c.id) as num_meetings,
                   (SELECT COUNT(*) FROM meetings WHERE client_id = c.id AND status = 'done') as analyzed_meetings
            FROM clients c ORDER BY c.name
            """
        ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "고객 목록"

    # 헤더
    headers = ["ID", "이름", "아파트/단지", "전화번호", "이메일", "방문경로", "단계",
               "첫 방문", "마지막 상담", "상담 건수", "분석 완료"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 데이터
    for row_idx, c in enumerate(clients, 2):
        ws.cell(row=row_idx, column=1, value=c["id"])
        ws.cell(row=row_idx, column=2, value=c["name"])
        ws.cell(row=row_idx, column=3, value=c["descriptor"] or "")
        ws.cell(row=row_idx, column=4, value=c["phone"] or "")
        ws.cell(row=row_idx, column=5, value=c["email"] or "")
        ws.cell(row=row_idx, column=6, value=c["visit_source"] or "")
        ws.cell(row=row_idx, column=7, value=c["stage"] or "초도")
        ws.cell(row=row_idx, column=8, value=(c["first_met_at"] or "")[:10])
        ws.cell(row=row_idx, column=9, value=(c["last_meeting_at"] or "")[:10])
        ws.cell(row=row_idx, column=10, value=c["num_meetings"])
        ws.cell(row=row_idx, column=11, value=c["analyzed_meetings"])

    # 열 너비 자동 조정
    col_widths = [8, 15, 25, 15, 25, 12, 8, 14, 14, 10, 10]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"InterioNote_고객목록_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.get("/api/export/meetings.xlsx")
def export_meetings():
    """상담 전체 목록을 xlsx 로 내보내기."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl 라이브러리가 없습니다.")

    with db_cursor() as cur:
        meetings = cur.execute(
            """
            SELECT m.id, c.name as client_name, c.descriptor, m.meeting_type,
                   m.started_at, m.duration_sec, m.status,
                   m.contract_amount, m.deposit_amount, m.contract_date,
                   (SELECT COUNT(*) FROM transcript_segments WHERE meeting_id = m.id) as segments,
                   (SELECT 1 FROM analyses WHERE meeting_id = m.id LIMIT 1) as has_analysis
            FROM meetings m
            JOIN clients c ON c.id = m.client_id
            ORDER BY m.started_at DESC
            """
        ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "상담 목록"

    headers = ["ID", "고객명", "단지", "상담종류", "일시", "시간(분)", "상태",
               "계약금액(원)", "입금액(원)", "계약일", "카드수", "AI분석"]
    header_fill = PatternFill("solid", fgColor="70AD47")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, m in enumerate(meetings, 2):
        dur_min = round(m["duration_sec"] / 60, 1) if m["duration_sec"] else 0
        ws.cell(row=row_idx, column=1, value=m["id"])
        ws.cell(row=row_idx, column=2, value=m["client_name"])
        ws.cell(row=row_idx, column=3, value=m["descriptor"] or "")
        ws.cell(row=row_idx, column=4, value=m["meeting_type"])
        ws.cell(row=row_idx, column=5, value=(m["started_at"] or "")[:16].replace("T", " "))
        ws.cell(row=row_idx, column=6, value=dur_min)
        ws.cell(row=row_idx, column=7, value=m["status"])
        ws.cell(row=row_idx, column=8, value=m["contract_amount"] or "")
        ws.cell(row=row_idx, column=9, value=m["deposit_amount"] or "")
        ws.cell(row=row_idx, column=10, value=m["contract_date"] or "")
        ws.cell(row=row_idx, column=11, value=m["segments"])
        ws.cell(row=row_idx, column=12, value="✅" if m["has_analysis"] else "")

    col_widths = [8, 15, 20, 12, 18, 10, 10, 15, 12, 12, 8, 8]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"InterioNote_상담목록_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


# ─── ⑭ QR 코드 생성 ──────────────────────────────────────────────────────────

@router.get("/api/qr/{meeting_id}")
def generate_qr(meeting_id: int):
    """상담 요약 페이지 QR 코드 이미지 반환 (PNG).
    QR 이 가리키는 URL: /meetings/{id}/print-summary (로컬 네트워크에서 접근 가능 시 사용)
    """
    try:
        import qrcode
        from qrcode.image.pure import PyPNGImage
    except ImportError:
        # qrcode 없으면 SVG fallback (CDN JS로 클라이언트 생성 유도)
        raise HTTPException(501, "qrcode 라이브러리가 설치되지 않았습니다. pip install qrcode[pil]")

    with db_cursor() as cur:
        m = cur.execute("SELECT id FROM meetings WHERE id = ?", (meeting_id,)).fetchone()
        if not m:
            raise HTTPException(404, "상담을 찾을 수 없습니다.")

    url = f"http://localhost:8000/meetings/{meeting_id}/print-summary"
    qr = qrcode.QRCode(version=1, box_size=8, border=4)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)

    return Response(content=buf.getvalue(), media_type="image/png")
