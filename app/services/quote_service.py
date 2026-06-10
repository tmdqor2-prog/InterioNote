"""
v2.8.0 FF — 견적서 자동 작성 (사용자 양식 기반).

설계 철학:
- 사용자 견적서 양식 (xlsx) 은 디자이너마다 다름 → 매핑 마법사로 받음 (OJT 와 동일 패턴)
- 견적서의 핵심 셀 (고객정보 시트의 B3, B4, B5 등) 을 사용자가 알려줌
- AI 분석 결과의 site_info + meeting + client 정보로 자동 채움
- 채워진 사본은 고객 폴더의 상담기록/{날짜}/{type}_견적_초안.xlsx 로 저장
- 사용자가 엑셀에서 직접 열고 견적 항목 입력 (단가표 자동 계산은 v2.8.1 예정)

settings DB 키:
- quote.template_path     : 견적서 템플릿 xlsx 절대 경로
- quote.template_name     : 사용자 친화적 이름 (예: "전체견적서_신입용")
- quote.cell_mapping      : JSON {"client_name": {"sheet": "고객정보", "cell": "B3"}, ...}
- quote.designer_name     : 담당자 이름 (자동 채우기용 — 매번 같음)
"""
from __future__ import annotations

import json
import shutil
from datetime import datetime, date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from app import config
from app.db import db_cursor, get_setting, set_setting


# --------------------------------------------------------------------------
# 설정
# --------------------------------------------------------------------------
def get_config() -> dict:
    return {
        "template_path": get_setting("quote.template_path", "") or "",
        "template_name": get_setting("quote.template_name", "") or "",
        "cell_mapping": json.loads(get_setting("quote.cell_mapping", "{}") or "{}"),
        "designer_name": get_setting("quote.designer_name", "") or "",
    }


def save_config(*, template_path: str, template_name: str = "",
                cell_mapping: dict[str, Any] | None = None,
                designer_name: str = "") -> None:
    set_setting("quote.template_path", template_path or "")
    set_setting("quote.template_name", template_name or "")
    set_setting("quote.cell_mapping", json.dumps(cell_mapping or {}, ensure_ascii=False))
    set_setting("quote.designer_name", designer_name or "")


def is_configured() -> bool:
    cfg = get_config()
    return bool(cfg["template_path"] and cfg["cell_mapping"])


# --------------------------------------------------------------------------
# 템플릿 분석 (마법사용)
# --------------------------------------------------------------------------
def analyze_template(file_path: str) -> dict:
    """견적서 템플릿 파일의 시트 구조 + 첫 N 셀 미리보기.

    매핑 마법사 UI 가 쓰는 정보:
    - sheets: 시트 목록
    - sheets_preview: 각 시트의 처음 20×8 셀 미리보기 (어느 셀에 무슨 라벨이 있는지)
    """
    p = Path(file_path)
    if not p.exists():
        raise FileNotFoundError(f"견적서 템플릿을 찾을 수 없습니다: {file_path}")
    if p.stat().st_size < 4096:
        raise ValueError(
            f"파일이 너무 작습니다 ({p.stat().st_size}B). "
            "OneDrive 클라우드 전용 파일일 수 있습니다 — 우클릭 → '항상 이 장치에 보관' 으로 받은 후 다시 시도하세요."
        )
    try:
        wb = load_workbook(file_path, data_only=False, read_only=True, keep_vba=False)
    except Exception as e:
        raise ValueError(f"엑셀을 열 수 없습니다: {type(e).__name__}: {e}")

    out: dict[str, Any] = {
        "sheets": list(wb.sheetnames),
        "active_sheet": wb.active.title if wb.active else (wb.sheetnames[0] if wb.sheetnames else ""),
        "sheets_preview": {},
        "size_bytes": p.stat().st_size,
    }
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_col = min(ws.max_column or 1, 15)
        max_row = min(ws.max_row or 1, 30)
        rows = []
        for r_idx in range(1, max_row + 1):
            cells = []
            for c_idx in range(1, max_col + 1):
                v = ws.cell(row=r_idx, column=c_idx).value
                if isinstance(v, datetime):
                    v = v.isoformat(timespec="seconds")
                elif v is None:
                    v = ""
                else:
                    v = str(v)
                cells.append({
                    "row": r_idx, "col": c_idx,
                    "address": f"{_col_letter(c_idx)}{r_idx}",
                    "value": v[:40] if isinstance(v, str) else v,
                })
            rows.append(cells)
        out["sheets_preview"][sheet_name] = {
            "max_col": max_col,
            "max_row": max_row,
            "rows": rows,
        }
    wb.close()
    return out


def _col_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


# --------------------------------------------------------------------------
# 분석 결과 → 견적서 필드 데이터 추출
# --------------------------------------------------------------------------
def build_quote_data(meeting_id: int) -> dict[str, str]:
    """meeting_id + AI 분석 → 견적서 필드 dict (QUOTE_FIELD_KEYS 기준)."""
    cfg = get_config()
    designer_name = cfg.get("designer_name", "") or ""

    with db_cursor() as cur:
        m = cur.execute(
            """
            SELECT m.id, m.meeting_type, m.started_at, m.notes,
                   c.name AS client_name, c.descriptor, c.folder_name, c.folder_path,
                   c.phone AS client_phone
            FROM meetings m
            JOIN clients c ON c.id = m.client_id
            WHERE m.id = ?
            """,
            (meeting_id,),
        ).fetchone()
        if m is None:
            raise ValueError(f"meeting {meeting_id} not found")
        analysis_row = cur.execute(
            "SELECT data_json FROM analyses WHERE meeting_id = ?", (meeting_id,)
        ).fetchone()
        analysis = json.loads(analysis_row["data_json"]) if analysis_row else {}

    site = analysis.get("site_info") or {}

    # 주소 합성
    addr_parts = []
    for k in ("address", "complex_name", "building_unit"):
        v = site.get(k)
        if v:
            addr_parts.append(str(v).strip())
    site_address = " ".join(addr_parts) or (m["descriptor"] or "")

    # 평형
    size_pyeong = ""
    if site.get("size_pyeong"):
        size_pyeong = str(site["size_pyeong"])
    size_m2 = ""
    if site.get("size_m2"):
        size_m2 = str(site["size_m2"])

    # 공사 종류 — client_requests(고객 요청사항 목록) 우선 → structure_type
    construction_type = ""
    req = analysis.get("client_requests") or []
    if isinstance(req, list) and req:
        construction_type = ", ".join(str(r).strip() for r in req if r)[:100]
    elif isinstance(req, str) and req.strip():
        construction_type = req.strip()[:100]
    if not construction_type:
        construction_type = (site.get("structure_type") or "").strip()

    # 전화번호 — clients.phone DB 에서 자동 연동
    client_phone = (m["client_phone"] or "").strip()

    # 견적일 = today
    today = date.today()

    return {
        "client_name": (m["client_name"] or "").strip(),
        "client_phone": client_phone,  # clients.phone 에서 자동 연동
        "site_address": site_address.strip(),
        "construction_period": "",
        "designer_name": designer_name,
        "size_pyeong": size_pyeong,
        "size_m2": size_m2,
        "quote_date": today.strftime("%Y. %m. %d"),
        "construction_type": construction_type,
        "memo": (analysis.get("summary") or "").strip()[:300],
    }


# --------------------------------------------------------------------------
# 견적서 초안 생성 (템플릿 복사 + 셀 자동 채우기)
# --------------------------------------------------------------------------
def generate_quote_draft(meeting_id: int, *, override_data: dict | None = None) -> dict:
    """견적서 템플릿을 고객 폴더로 복사 + 매핑된 셀 자동 채우기.

    저장 위치: {meeting_folder}/{meeting_type}_견적초안_{YYYYMMDD_HHMMSS}.xlsx
    """
    cfg = get_config()
    template_path = cfg["template_path"]
    if not template_path:
        raise ValueError("견적서 템플릿이 설정되지 않았습니다. 설정 → 견적서 양식 에서 등록하세요.")
    tpl = Path(template_path)
    if not tpl.exists():
        raise FileNotFoundError(
            f"견적서 템플릿을 찾을 수 없습니다: {template_path}\n"
            "OneDrive 클라우드 전용 파일이라면 '항상 이 장치에 보관' 으로 받아주세요."
        )

    mapping = cfg["cell_mapping"] or {}
    if not mapping:
        raise ValueError("셀 매핑이 설정되지 않았습니다. 설정 → 견적서 양식 에서 매핑 마법사를 진행하세요.")

    data = override_data if override_data is not None else build_quote_data(meeting_id)

    # 저장 위치 — meeting_folder 가 있으면 거기, 없으면 고객 폴더
    with db_cursor() as cur:
        m = cur.execute(
            """
            SELECT m.meeting_type, m.meeting_folder, c.folder_path
            FROM meetings m JOIN clients c ON c.id = m.client_id
            WHERE m.id = ?
            """,
            (meeting_id,),
        ).fetchone()
        if m is None:
            raise ValueError(f"meeting {meeting_id} not found")

    target_dir = Path(m["meeting_folder"]) if m["meeting_folder"] else Path(m["folder_path"])
    target_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = target_dir / f"{m['meeting_type']}_견적초안_{ts}.xlsx"

    # 1) 템플릿 복사
    try:
        shutil.copy2(template_path, output_path)
    except Exception as e:
        raise RuntimeError(f"템플릿 복사 실패: {type(e).__name__}: {e}")

    # 2) 복사본 열어서 셀 채우기
    try:
        wb = load_workbook(output_path, keep_vba=False)
    except PermissionError as e:
        raise RuntimeError(
            "복사본을 열 수 없습니다 — 엑셀에서 출력 파일이 열려 있을 가능성. 닫고 다시 시도해주세요."
        ) from e
    except Exception as e:
        # 복사본 정리
        try: output_path.unlink()
        except Exception: pass
        raise RuntimeError(f"엑셀 처리 실패: {type(e).__name__}: {e}") from e

    written: list[dict[str, str]] = []
    try:
        for field_key, target in mapping.items():
            if not isinstance(target, dict):
                continue
            sheet_name = (target.get("sheet") or "").strip()
            cell_addr = (target.get("cell") or "").strip()
            if not sheet_name or not cell_addr:
                continue
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            value = (data.get(field_key) or "").strip()
            if not value:
                continue
            try:
                ws[cell_addr] = value
                written.append({"field": field_key, "sheet": sheet_name, "cell": cell_addr, "value": value[:60]})
            except Exception as e:
                # 잘못된 셀 주소는 건너뜀
                print(f"[quote] cell {sheet_name}!{cell_addr} write 실패: {e}", flush=True)
    except Exception as e:
        wb.close()
        raise RuntimeError(f"셀 쓰기 중 오류: {type(e).__name__}: {e}") from e

    # 3) 저장
    try:
        wb.save(output_path)
    except PermissionError as e:
        raise RuntimeError(
            "견적서 초안 저장 실패 — 출력 파일이 이미 엑셀에서 열려있는지 확인하세요."
        ) from e
    finally:
        try:
            wb.close()
        except Exception:
            pass

    return {
        "ok": True,
        "output_path": str(output_path),
        "written_cells": written,
        "written_count": len(written),
        "size_bytes": output_path.stat().st_size,
    }
