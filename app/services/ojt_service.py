"""
v2.6.0 JJ — OJT 엑셀 자동 동기화.

핵심 설계 원칙:
1. 디자이너마다 OJT 양식이 다름 → 매핑이 핵심.
2. 사용자가 자기 OJT 파일 + 컬럼 매핑 (마법사) 한 번 등록.
3. 신규 고객 초도상담 종료 후 → AI 분석에서 데이터 추출 → OJT 에 새 행 추가.
4. **신규 고객 초도상담 전용** (사용자 명시) — 디자인/견적 미팅은 스킵.
5. 충돌 방지: 파일 lock 시 친절한 에러. 백업 자동 생성.

Settings DB 키:
- ojt.file_path        : OJT 엑셀 파일 절대 경로 (str)
- ojt.column_mapping   : JSON {"name": "C", "phone": "D", ...} 또는 인덱스 {"name": 3, ...}
- ojt.start_row        : 데이터 시작 행 번호 (헤더 다음 행, 1-based)
- ojt.sheet_name       : 시트 이름 (None 이면 첫 번째 시트)
"""
from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from app import config
from app.db import db_cursor, get_setting, set_setting


# --------------------------------------------------------------------------
# 설정 헬퍼
# --------------------------------------------------------------------------
def get_config() -> dict:
    """현재 OJT 동기화 설정. 없으면 빈 dict."""
    return {
        "file_path": get_setting("ojt.file_path", "") or "",
        "column_mapping": json.loads(get_setting("ojt.column_mapping", "{}") or "{}"),
        "start_row": int(get_setting("ojt.start_row", "2") or 2),
        "sheet_name": get_setting("ojt.sheet_name", "") or None,
    }


def save_config(file_path: str, column_mapping: dict[str, Any],
                start_row: int = 2, sheet_name: str | None = None) -> None:
    set_setting("ojt.file_path", file_path or "")
    set_setting("ojt.column_mapping", json.dumps(column_mapping or {}, ensure_ascii=False))
    set_setting("ojt.start_row", str(int(start_row or 2)))
    set_setting("ojt.sheet_name", sheet_name or "")


def is_configured() -> bool:
    cfg = get_config()
    return bool(cfg["file_path"] and cfg["column_mapping"])


# --------------------------------------------------------------------------
# 파일 분석 (마법사)
# --------------------------------------------------------------------------
def analyze_file(file_path: str, *, head_rows: int = 5) -> dict:
    """OJT 파일 첫 N 행 읽고 시트/컬럼 정보 반환 (매핑 마법사용).

    반환:
    {
      "sheets": ["시트1", ...],
      "active_sheet": "시트1",
      "preview": {
          "시트1": {
              "max_col": 15,
              "rows": [
                  {"row": 1, "cells": [{"col": 1, "letter": "A", "value": "..."}, ...]},
                  ...
              ]
          },
          ...
      },
      "size_bytes": int,
    }
    """
    p = Path(file_path)
    if not p.exists():
        raise FileNotFoundError(f"OJT 파일을 찾을 수 없습니다: {file_path}")
    if p.stat().st_size < 4096:
        # OneDrive on-demand placeholder 의심 (실제 xlsx 는 최소 4KB+)
        raise ValueError(
            f"파일이 너무 작습니다 ({p.stat().st_size}B). "
            "OneDrive 클라우드 전용 파일일 수 있습니다 — 파일을 우클릭 → '항상 이 장치에 보관' 으로 다운로드 후 다시 시도하세요."
        )

    try:
        wb = load_workbook(file_path, data_only=True, read_only=True, keep_vba=False)
    except Exception as e:
        raise ValueError(f"엑셀 파일을 열 수 없습니다: {type(e).__name__}: {e}")

    out: dict[str, Any] = {
        "sheets": list(wb.sheetnames),
        "active_sheet": wb.active.title if wb.active else (wb.sheetnames[0] if wb.sheetnames else ""),
        "preview": {},
        "size_bytes": p.stat().st_size,
    }
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_col = min(ws.max_column or 1, 30)  # 30 컬럼 이상은 미리보기에서 자름
        rows_preview = []
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=head_rows, max_col=max_col, values_only=False), start=1):
            cells = []
            for cell in row:
                v = cell.value
                if isinstance(v, datetime):
                    v = v.isoformat(timespec="seconds")
                elif v is None:
                    v = ""
                else:
                    v = str(v)
                cells.append({
                    "col": cell.column,
                    "letter": cell.column_letter,
                    "value": v[:60] if isinstance(v, str) else v,
                })
            rows_preview.append({"row": r_idx, "cells": cells})
        out["preview"][sheet_name] = {"max_col": max_col, "rows": rows_preview}
    wb.close()
    return out


# --------------------------------------------------------------------------
# 동기화 — 데이터 추출
# --------------------------------------------------------------------------
def build_row_data(meeting_id: int) -> dict[str, str]:
    """초도상담 + AI 분석 결과 → OJT 필드 dict.

    추출 대상 (OJT_FIELD_KEYS):
      consult_date, name, phone, address, construction, size, budget,
      measure_date, probability, memo
    """
    with db_cursor() as cur:
        meeting = cur.execute(
            """
            SELECT m.id, m.meeting_type, m.started_at, m.ended_at, m.duration_sec, m.notes,
                   c.name AS client_name, c.descriptor, c.folder_name,
                   c.phone AS client_phone
            FROM meetings m
            JOIN clients c ON c.id = m.client_id
            WHERE m.id = ?
            """,
            (meeting_id,),
        ).fetchone()
        if meeting is None:
            raise ValueError(f"meeting {meeting_id} not found")
        analysis_row = cur.execute(
            "SELECT data_json FROM analyses WHERE meeting_id = ?", (meeting_id,)
        ).fetchone()
        analysis_data = json.loads(analysis_row["data_json"]) if analysis_row else {}

    site_info = analysis_data.get("site_info") or {}
    client_info = analysis_data.get("client_info") or {}
    summary = (analysis_data.get("summary") or "").strip()

    # 상담 일자 (started_at 의 날짜만)
    consult_date = ""
    try:
        sa = meeting["started_at"]
        if sa:
            d = datetime.fromisoformat(sa)
            consult_date = d.strftime("%Y-%m-%d")
    except Exception:
        pass

    # 주소 합성
    addr_parts = []
    if site_info.get("address"):
        addr_parts.append(str(site_info["address"]).strip())
    if site_info.get("complex_name"):
        addr_parts.append(str(site_info["complex_name"]).strip())
    if site_info.get("building_unit"):
        addr_parts.append(str(site_info["building_unit"]).strip())
    address = " ".join(p for p in addr_parts if p) or (meeting["descriptor"] or "")

    # 평형
    size_parts = []
    if site_info.get("size_pyeong"):
        size_parts.append(f"{site_info['size_pyeong']}평")
    if site_info.get("size_m2"):
        size_parts.append(f"{site_info['size_m2']}m²")
    if site_info.get("expansion"):
        size_parts.append("확장")
    size_str = " ".join(size_parts)

    # 공사 내용 — client_requests(고객 요청사항 목록) 우선 → structure_type → 요약 첫줄
    construction = ""
    req = analysis_data.get("client_requests") or []
    if isinstance(req, list) and req:
        construction = ", ".join(str(r).strip() for r in req if r)[:200]
    elif isinstance(req, str) and req.strip():
        construction = req.strip()[:200]
    if not construction:
        construction = (site_info.get("structure_type") or "").strip()
    if not construction and summary:
        construction = summary.split("\n")[0][:200]

    # 예산 — client_info.budget / budget_range / analysis 최상위 budget 순서로 시도
    budget = ""
    for k in ("budget", "budget_range", "예산"):
        for src in (client_info, analysis_data, site_info):
            v = src.get(k) if isinstance(src, dict) else None
            if v:
                budget = str(v)
                break
        if budget:
            break

    # 전화번호 — clients.phone DB 에서 자동 연동
    phone = (meeting["client_phone"] or "").strip()

    # 메모 = 사용자 메모 + AI 요약 첫 줄
    memo_parts = []
    if meeting["notes"]:
        memo_parts.append(meeting["notes"].strip()[:200])
    if summary:
        memo_parts.append(summary.split("\n")[0][:200])
    memo = " | ".join(memo_parts)

    return {
        "consult_date": consult_date,
        "name": (meeting["client_name"] or "").strip(),
        "phone": phone,  # clients.phone 에서 자동 연동
        "address": address.strip(),
        "construction": construction.strip()[:200],
        "size": size_str,
        "budget": budget,
        "measure_date": "",  # 분석에서 추출 안 됨
        "probability": "B",  # 기본값 (애매) — 사용자가 동기화 모달에서 수정
        "memo": memo,
    }


# --------------------------------------------------------------------------
# 동기화 — 엑셀 쓰기
# --------------------------------------------------------------------------
def _resolve_col(col_spec: Any) -> int:
    """컬럼 매핑 값을 1-based 정수 인덱스로. 'C' 같은 문자도 허용."""
    if isinstance(col_spec, int):
        return max(1, col_spec)
    if isinstance(col_spec, str):
        s = col_spec.strip()
        if not s:
            return 0
        if s.isdigit():
            return max(1, int(s))
        try:
            return column_index_from_string(s.upper())
        except Exception:
            return 0
    return 0


def append_row_to_ojt(meeting_id: int, *, override_data: dict | None = None) -> dict:
    """OJT 파일의 다음 빈 행에 데이터 한 줄 추가.

    override_data: 사용자가 동기화 모달에서 직접 편집한 값 (없으면 build_row_data 결과 그대로).
    """
    cfg = get_config()
    file_path = cfg["file_path"]
    if not file_path:
        raise ValueError("OJT 파일 경로가 설정되지 않았습니다. 설정 → OJT 연동 에서 등록하세요.")
    p = Path(file_path)
    if not p.exists():
        raise FileNotFoundError(
            f"OJT 파일을 찾을 수 없습니다: {file_path}\n"
            "OneDrive 클라우드 전용 파일이라면 우클릭 → '항상 이 장치에 보관' 으로 다운로드 후 다시 시도하세요."
        )

    mapping = cfg["column_mapping"] or {}
    if not mapping:
        raise ValueError("OJT 컬럼 매핑이 설정되지 않았습니다. 설정 → OJT 연동 에서 매핑하세요.")

    data = override_data if override_data is not None else build_row_data(meeting_id)

    # 백업 자동 생성 (~InterioNote_OJT_backup_YYYYMMDD_HHMMSS.xlsx)
    try:
        backup_dir = config.USER_DATA_DIR / "ojt_backup"
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / f"{p.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        shutil.copy2(p, backup_path)
    except PermissionError as e:
        raise RuntimeError(
            "백업 생성 실패 — OJT 파일이 다른 곳에서 열려 있을 가능성. "
            "엑셀·다른 프로그램에서 닫고 다시 시도해주세요."
        ) from e
    except Exception as e:
        # 백업 실패해도 계속 진행 (자동 백업은 보너스)
        print(f"[ojt_sync] 백업 생성 실패 (무시): {type(e).__name__}: {e}", flush=True)
        backup_path = p  # 백업 없음 표시용

    try:
        wb = load_workbook(file_path, keep_vba=False)
    except PermissionError as e:
        raise RuntimeError(
            "OJT 파일을 열 수 없습니다 — 엑셀에서 OJT 파일이 열려 있을 가능성이 높습니다. "
            "엑셀에서 파일을 닫고 다시 시도해주세요."
        ) from e
    except Exception as e:
        raise RuntimeError(
            f"OJT 파일을 열 수 없습니다 ({type(e).__name__}). "
            "OneDrive 클라우드 전용 / 파일 손상 / 엑셀에서 열림 여부를 확인하세요."
        ) from e

    sheet_name = cfg["sheet_name"]
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # 매핑된 컬럼 검증
    mapped_cols = [c for c in (_resolve_col(v) for v in mapping.values()) if c > 0]
    if not mapped_cols:
        wb.close()
        raise ValueError(
            "매핑된 컬럼이 없습니다. 설정 → OJT 연동 에서 다시 매핑해주세요. "
            f"(현재 매핑: {mapping})"
        )

    # 다음 빈 행 찾기 — 고객명(name) 컬럼 기준으로만 체크
    # 이유: 다른 컬럼에 수식·기본값이 있어도 고객명이 빈 행이면 미사용 행
    name_col = _resolve_col(mapping.get("name", ""))
    check_col = name_col if name_col > 0 else (min(mapped_cols) if mapped_cols else 1)

    start_row = max(1, cfg["start_row"])
    target_row = start_row
    while True:
        if target_row > 9999:
            wb.close()
            raise RuntimeError(
                "OJT 파일에 빈 행을 찾을 수 없습니다 (9999행까지 모두 채워짐). "
                "엑셀에서 OJT 파일을 직접 정리해주세요."
            )
        if ws.cell(row=target_row, column=check_col).value in (None, ""):
            break
        target_row += 1

    # 행 쓰기
    written: dict[str, str] = {}
    try:
        for field_key, col_spec in mapping.items():
            col = _resolve_col(col_spec)
            if col <= 0:
                continue
            value = (data.get(field_key) or "").strip()
            if value:
                ws.cell(row=target_row, column=col).value = value
                written[field_key] = value
    except Exception as e:
        wb.close()
        raise RuntimeError(
            f"OJT 행 데이터 쓰기 중 오류 ({type(e).__name__}: {e}). "
            "매핑이 올바른지 설정에서 확인해주세요."
        ) from e

    # 저장 — PermissionError 가 가장 흔한 실패 (엑셀에서 파일 열림)
    try:
        wb.save(file_path)
    except PermissionError as e:
        # OneDrive 동기화 잠금 또는 엑셀에서 파일 열림
        raise RuntimeError(
            "OJT 파일에 저장할 수 없습니다 — 다음 중 하나일 가능성:\n"
            "  ① 엑셀에서 OJT 파일이 열려 있음 → 엑셀에서 파일을 닫아주세요\n"
            "  ② OneDrive 동기화 중 → 잠시 기다린 후 다시 시도\n"
            "  ③ 파일이 읽기 전용 → 파일 속성에서 '읽기 전용' 해제\n"
            f"({type(e).__name__})"
        ) from e
    except Exception as e:
        raise RuntimeError(
            f"OJT 저장 실패 ({type(e).__name__}: {str(e)[:200]})"
        ) from e
    finally:
        try:
            wb.close()
        except Exception:
            pass

    # DB 에 동기화 기록
    try:
        with db_cursor() as cur:
            cur.execute(
                "UPDATE meetings SET ojt_synced_at = ? WHERE id = ?",
                (datetime.now().isoformat(timespec="seconds"), meeting_id),
            )
    except Exception as e:
        # DB 기록 실패해도 OJT 자체는 추가됨 — 경고만
        print(f"[ojt_sync] DB ojt_synced_at 갱신 실패 (무시): {type(e).__name__}: {e}", flush=True)

    return {
        "ok": True,
        "row": target_row,
        "written_fields": written,
        "backup_path": str(backup_path) if backup_path != p else "",
    }
